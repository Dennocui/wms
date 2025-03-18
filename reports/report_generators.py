import os
import pandas as pd
import numpy as np
import csv
import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Sum, Avg, F, Q
from django.core.files.base import ContentFile
from django.conf import settings
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.contrib.auth.models import User
from users.models import Activity, UserProfile
from warehouse.models import Warehouse
from inventory.models import Product, StockMovement, Category
from orders.models import Order, OrderItem

def generate_report(generated_report):
    """
    Main function to generate a report based on the report type
    """
    report = generated_report.report
    parameters = generated_report.parameters_used or {}
    
    try:
        # Call the appropriate report generator based on report type
        if report.report_type == 'INVENTORY':
            file_content = generate_inventory_report(report, parameters)
        elif report.report_type == 'ORDER':
            file_content = generate_order_report(report, parameters)
        elif report.report_type == 'SHIPPING':
            file_content = generate_shipping_report(report, parameters)
        elif report.report_type == 'RECEIVING':
            file_content = generate_receiving_report(report, parameters)
        elif report.report_type == 'USER_ACTIVITY':
            file_content = generate_user_activity_report(report, parameters)
        elif report.report_type == 'PERFORMANCE':
            file_content = generate_performance_report(report, parameters)
        elif report.report_type == 'SALES':
            file_content = generate_sales_report(report, parameters)
        elif report.report_type == 'CUSTOM':
            file_content = generate_custom_report(report, parameters)
        else:
            raise ValueError(f"Unknown report type: {report.report_type}")
        
        # Save the generated file
        filename = f"{report.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{report.format.lower()}"
        generated_report.file.save(filename, ContentFile(file_content))
        
        # Update the generated report status
        generated_report.status = 'COMPLETED'
        generated_report.end_time = timezone.now()
        generated_report.save()
        
        return generated_report
    
    except Exception as e:
        # Update the generated report with error info
        generated_report.status = 'FAILED'
        generated_report.error_message = str(e)
        generated_report.end_time = timezone.now()
        generated_report.save()
        raise

def generate_inventory_report(report, parameters):
    """
    Generate an inventory report
    """
    # Get products based on parameters
    products = Product.objects.all()
    
    # Apply filters from parameters
    if 'category_id' in parameters:
        products = products.filter(category_id=parameters['category_id'])
    
    if 'warehouse_id' in parameters:
        products = products.filter(warehouse_id=parameters['warehouse_id'])
    
    if 'min_stock' in parameters:
        products = products.filter(current_stock__gte=parameters['min_stock'])
    
    if 'max_stock' in parameters:
        products = products.filter(current_stock__lte=parameters['max_stock'])
    
    # Convert queryset to DataFrame
    data = list(products.values(
        'id', 'name', 'sku', 'category__name', 'current_stock', 
        'min_stock_level', 'max_stock_level', 'reorder_point',
        'unit_price', 'cost_price', 'warehouse__name'
    ))
    
    df = pd.DataFrame(data)
    
    # Generate report based on format
    if report.format == 'PDF':
        return generate_pdf_inventory_report(df, report, parameters)
    elif report.format == 'CSV':
        return generate_csv_inventory_report(df, report, parameters)
    elif report.format == 'EXCEL':
        return generate_excel_inventory_report(df, report, parameters)
    elif report.format == 'JSON':
        return generate_json_inventory_report(df, report, parameters)
    else:
        raise ValueError(f"Unsupported format: {report.format}")

def generate_order_report(report, parameters):
    """
    Generate an order report
    """
    # Get orders based on parameters
    orders = Order.objects.all()
    
    # Apply filters from parameters
    if 'start_date' in parameters:
        start_date = datetime.strptime(parameters['start_date'], '%Y-%m-%d')
        orders = orders.filter(created_at__gte=start_date)
    
    if 'end_date' in parameters:
        end_date = datetime.strptime(parameters['end_date'], '%Y-%m-%d')
        orders = orders.filter(created_at__lte=end_date)
    
    if 'status' in parameters:
        orders = orders.filter(status=parameters['status'])
    
    if 'warehouse_id' in parameters:
        orders = orders.filter(warehouse_id=parameters['warehouse_id'])
    
    # Convert queryset to DataFrame
    data = list(orders.values(
        'id', 'order_number', 'customer_name', 'status', 'total_amount',
        'created_at', 'completed_at', 'warehouse__name'
    ))
    
    df = pd.DataFrame(data)
    
    # Generate report based on format
    if report.format == 'PDF':
        return generate_pdf_order_report(df, report, parameters)
    elif report.format == 'CSV':
        return generate_csv_order_report(df, report, parameters)
    elif report.format == 'EXCEL':
        return generate_excel_order_report(df, report, parameters)
    elif report.format == 'JSON':
        return generate_json_order_report(df, report, parameters)
    else:
        raise ValueError(f"Unsupported format: {report.format}")

# Additional report generator functions would be implemented here for each report type

def generate_pdf_inventory_report(df, report, parameters):
    """
    Generate a PDF inventory report
    """
    # Create a BytesIO object to store the PDF
    buffer = BytesIO()
    
    # Create a new PDF with Matplotlib
    with PdfPages(buffer) as pdf:
        # Create a figure and axis
        plt.figure(figsize=(10, 6))
        
        # Create a title
        plt.suptitle(f"Inventory Report - {datetime.now().strftime('%Y-%m-%d')}", fontsize=16)
        
        # Create a table
        table_data = df.values
        col_labels = df.columns
        
        # Create the table
        plt.table(cellText=table_data, colLabels=col_labels, loc='center')
        
        # Hide axes
        plt.axis('off')
        
        # Save the figure
        pdf.savefig()
        plt.close()
        
        # Add a page with inventory statistics
        plt.figure(figsize=(10, 6))
        
        # Create a title
        plt.suptitle(f"Inventory Statistics - {datetime.now().strftime('%Y-%m-%d')}", fontsize=16)
        
        # Calculate statistics
        total_products = len(df)
        total_value = (df['current_stock'] * df['unit_price']).sum()
        low_stock = len(df[df['current_stock'] < df['reorder_point']])
        
        # Create a table with statistics
        stats_data = [
            ['Total Products', total_products],
            ['Total Inventory Value', f"${total_value:.2f}"],
            ['Low Stock Items', low_stock]
        ]
        
        # Create the table
        plt.table(cellText=stats_data, loc='center')
        
        # Hide axes
        plt.axis('off')
        
        # Save the figure
        pdf.savefig()
        plt.close()
    
    # Get the PDF content
    buffer.seek(0)
    return buffer.read()

def generate_csv_inventory_report(df, report, parameters):
    """
    Generate a CSV inventory report
    """
    # Create a BytesIO object to store the CSV
    buffer = BytesIO()
    
    # Write the DataFrame to the CSV
    df.to_csv(buffer, index=False)
    
    # Get the CSV content
    buffer.seek(0)
    return buffer.read()

def generate_excel_inventory_report(df, report, parameters):
    """
    Generate an Excel inventory report
    """
    # Create a BytesIO object to store the Excel file
    buffer = BytesIO()
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Add headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data
    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
    
    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Save the workbook
    wb.save(buffer)
    
    # Get the Excel content
    buffer.seek(0)
    return buffer.read()

def generate_json_inventory_report(df, report, parameters):
    """
    Generate a JSON inventory report
    """
    # Convert DataFrame to JSON
    json_data = df.to_json(orient='records')
    
    # Return the JSON as bytes
    return json_data.encode('utf-8')

# Similar functions would be implemented for other report types and formats